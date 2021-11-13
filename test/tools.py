from openpyxl import Workbook,load_workbook
from openpyxl.styles import *
import os
import warnings 
warnings.filterwarnings('ignore')
wb = load_workbook("test.xlsx")
ws = wb.active

for i in range (2,len(ws['A'])):
    ws[f'B{i}'] = f'MRBTS-{i}'
    #print(ws[f'B{i}'].value)
wb.save('test.xlsx')